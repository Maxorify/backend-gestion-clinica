"""
Router para la administración profesional de asistencia de doctores.
Endpoints: Marcas, Justificaciones, Reportes, Configuración.

Autor: Sistema de Gestión Clínica
Fecha: 2025
"""
from fastapi import APIRouter, HTTPException, Depends, Query, Request
from typing import List, Optional
from datetime import datetime, date, time, timedelta
from src.utils.supabase import supabase_client
from src.models.asistencia import (
    MarcaAsistenciaCreate, MarcaAsistenciaResponse,
    JustificacionCreate, EstadoAsistenciaResponse,
    TurnoAsistenciaDetalle, ResumenDiarioAsistencia,
    EstadisticasAsistenciaDoctor, DoctorBasicInfo,
    ParametroAsistencia, ParametroAsistenciaUpdate,
    TipoMarca, FuenteMarca, EstadoAsistencia
)

attendance_router = APIRouter(tags=["Asistencia Profesional"], prefix="/asistencia")


# ============================================================================
# ENDPOINT 1: RESUMEN DEL DÍA (Vista principal Timeline)
# ============================================================================

@attendance_router.get("/turnos-dia", response_model=ResumenDiarioAsistencia)
async def obtener_turnos_dia(
    fecha: Optional[date] = Query(None, description="Fecha a consultar (default: hoy)"),
):
    """
    📊 Obtiene el resumen consolidado de asistencia del día basado en horarios_personal.
    
    **Retorna:**
    - KPIs: Total turnos, en turno, asistieron, atrasos, ausentes, justificados
    - Lista de turnos con estado calculado, marcas, y productividad
    
    **Usado en:** Vista principal Timeline (asistencia.jsx)
    """
    fecha_consulta = fecha or date.today()
    
    try:
        # 1. Obtener horarios del día desde horarios_personal
        horarios_response = supabase_client.from_("horarios_personal") \
            .select("*, usuario:usuario_sistema_id(id, nombre, apellido_paterno, apellido_materno, rut)") \
            .gte("inicio_bloque", f"{fecha_consulta}T00:00:00") \
            .lte("inicio_bloque", f"{fecha_consulta}T23:59:59") \
            .order("inicio_bloque", desc=False) \
            .execute()
        
        if not horarios_response.data:
            return ResumenDiarioAsistencia(
                fecha=fecha_consulta,
                total_turnos=0,
                en_turno=0,
                asistieron=0,
                con_atraso=0,
                ausentes=0,
                justificados=0,
                turnos=[]
            )
        
        # 2. Agrupar horarios por doctor (cada doctor tiene múltiples bloques de 30 min)
        doctores_dict = {}
        for horario in horarios_response.data:
            if not horario.get('usuario'):
                continue
                
            doctor_id = horario['usuario']['id']
            
            if doctor_id not in doctores_dict:
                doctores_dict[doctor_id] = {
                    'doctor': horario['usuario'],
                    'bloques': [],
                    'inicio_turno': horario['inicio_bloque'],
                    'finalizacion_turno': horario['finalizacion_bloque']
                }
            else:
                # Actualizar fin de turno al último bloque
                doctores_dict[doctor_id]['finalizacion_turno'] = horario['finalizacion_bloque']
            
            doctores_dict[doctor_id]['bloques'].append(horario)
        
        # 3. Procesar cada doctor y obtener su asistencia
        turnos_procesados = []
        stats = {
            "en_turno": 0,
            "asistieron": 0,
            "con_atraso": 0,
            "ausentes": 0,
            "justificados": 0
        }
        
        for doctor_id, info in doctores_dict.items():
                turnos=[]
            )
        
        # Procesar turnos
        turnos_procesados = []
        stats = {
            "en_turno": 0,
            "asistieron": 0,
            "con_atraso": 0,
            "ausentes": 0,
            "justificados": 0
        }
        
        for turno in response.data:
            doctor_id = turno["doctor_id"]
            
            # Obtener datos completos del doctor
            doctor_response = supabase_client.from_("usuario_sistema") \
                .select("nombre, apellido_paterno, apellido_materno, rut") \
                .eq("id", doctor_id) \
                .single() \
                .execute()
            
            doctor_data = doctor_response.data if doctor_response.data else {}
            
            # Obtener especialidades del doctor
            esp_response = supabase_client.from_("especialidades_doctor") \
                .select("especialidad:especialidad_id(nombre)") \
                .eq("usuario_sistema_id", doctor_id) \
                .execute()
            
            especialidades = []
            if esp_response.data:
                especialidades = [e["especialidad"]["nombre"] for e in esp_response.data if e.get("especialidad")]
            
            # Construir objeto doctor
            doctor = DoctorBasicInfo(
                id=doctor_id,
                nombre=doctor_data.get("nombre", ""),
                apellido_paterno=doctor_data.get("apellido_paterno", ""),
                apellido_materno=doctor_data.get("apellido_materno", ""),
                nombre_completo=turno.get("doctor_nombre", ""),
                rut=doctor_data.get("rut"),
                especialidades=especialidades
            )
            
            # Obtener conteo de pacientes
            pacientes_response = supabase_client.from_("cita_medica") \
                .select("id", count="exact") \
                .eq("doctor_id", doctor_id) \
                .gte("fecha_atencion", f"{fecha_consulta}T00:00:00") \
                .lte("fecha_atencion", f"{fecha_consulta}T23:59:59") \
                .execute()
            
            # Para obtener pacientes atendidos, necesitamos consultar la tabla estado
            atendidos_response = supabase_client.from_("estado") \
                .select("id, cita_medica:cita_medica_id(doctor_id, fecha_atencion)", count="exact") \
                .eq("estado", "COMPLETADA") \
                .execute()
            
            # Filtrar por doctor y fecha del turno
            pacientes_atendidos = 0
            if atendidos_response.data:
                for est in atendidos_response.data:
                    cita = est.get("cita_medica", {})
                    if cita and cita.get("doctor_id") == doctor_id:
                        fecha_cita = cita.get("fecha_atencion", "")
                        if fecha_cita.startswith(str(fecha_consulta)):
                            pacientes_atendidos += 1
            
            pacientes_agendados = pacientes_response.count or 0
            
            # Estado de asistencia (default EN_TURNO si es None)
            estado = turno.get("estado") or "EN_TURNO"
            
            # Minutos de atraso (default 0 si es None)
            minutos_atraso = turno.get("minutos_atraso") or 0
            
            # Actualizar estadísticas
            if estado == "EN_TURNO":
                stats["en_turno"] += 1
            elif estado == "ASISTIO":
                stats["asistieron"] += 1
            elif estado == "ATRASO":
                stats["con_atraso"] += 1
            elif estado == "AUSENTE":
                stats["ausentes"] += 1
            elif estado == "JUSTIFICADO":
                stats["justificados"] += 1
            
            # Construir turno detallado
            turno_detalle = TurnoAsistenciaDetalle(
                id=turno["asistencia_id"],
                horario_id=None,  # La vista no incluye horario_id
                inicio_turno=datetime.fromisoformat(turno["inicio_turno"].replace("Z", "+00:00")),
                finalizacion_turno=datetime.fromisoformat(turno["finalizacion_turno"].replace("Z", "+00:00")) if turno.get("finalizacion_turno") else None,
                doctor=doctor,
                estado_asistencia=estado,
                minutos_atraso=minutos_atraso,
                minutos_trabajados=turno.get("minutos_trabajados"),
                porcentaje_asistencia=turno.get("porcentaje_asistencia"),
                marca_entrada=datetime.fromisoformat(turno["hora_entrada_real"].replace("Z", "+00:00")) if turno.get("hora_entrada_real") else None,
                marca_salida=datetime.fromisoformat(turno["hora_salida_real"].replace("Z", "+00:00")) if turno.get("hora_salida_real") else None,
                fuente_entrada=None,  # La vista no incluye fuentes
                fuente_salida=None,
                justificacion=turno.get("justificacion"),
                tipo_justificacion=turno.get("tipo_justificacion"),
                pacientes_agendados=pacientes_agendados,
                pacientes_atendidos=pacientes_atendidos,
                created_at=datetime.now()  # La vista no incluye created_at, usar fecha actual
            )
            
            turnos_procesados.append(turno_detalle)
        
        return ResumenDiarioAsistencia(
            fecha=fecha_consulta,
            total_turnos=len(turnos_procesados),
            en_turno=stats["en_turno"],
            asistieron=stats["asistieron"],
            con_atraso=stats["con_atraso"],
            ausentes=stats["ausentes"],
            justificados=stats["justificados"],
            turnos=turnos_procesados
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener turnos del día: {str(e)}")


# ============================================================================
# ENDPOINT 2: REGISTRAR MARCA MANUAL
# ============================================================================

@attendance_router.post("/marcar-manual", response_model=MarcaAsistenciaResponse)
async def registrar_marca_manual(
    marca: MarcaAsistenciaCreate,
    request: Request
):
    """
    ✍️ Registra una marca de entrada/salida de forma manual.
    
    **Casos de uso:**
    - Secretaria marca entrada/salida de doctor
    - Admin corrige marca errónea
    - Registro retroactivo autorizado
    
    **Automático:**
    - Calcula estado de asistencia después de insertar
    - Registra IP de origen
    - Audita quién hizo el registro
    """
    try:
        # Obtener IP del cliente
        client_ip = request.client.host if request.client else None
        
        # Preparar datos de la marca
        marca_data = {
            "usuario_sistema_id": marca.usuario_sistema_id,
            "horario_id": marca.horario_id,
            "tipo_marca": marca.tipo_marca.value,
            "fecha_hora_marca": marca.fecha_hora_marca.isoformat() if marca.fecha_hora_marca else datetime.now().isoformat(),
            "fuente": marca.fuente.value,
            "registrado_por": marca.registrado_por,
            "notas": marca.notas,
            "origen_ip": marca.origen_ip or client_ip
        }
        
        # Insertar marca
        response = supabase_client.from_("marcas_asistencia").insert(marca_data).execute()
        
        if not response.data:
            raise HTTPException(status_code=500, detail="No se pudo registrar la marca")
        
        marca_creada = response.data[0]
        
        # Buscar asistencia relacionada para recalcular estado
        asistencia_response = supabase_client.from_("asistencia") \
            .select("id") \
            .eq("usuario_sistema_id", marca.usuario_sistema_id) \
            .gte("inicio_turno", f"{marca_data['fecha_hora_marca'][:10]}T00:00:00") \
            .lte("inicio_turno", f"{marca_data['fecha_hora_marca'][:10]}T23:59:59") \
            .execute()
        
        if asistencia_response.data:
            # Ejecutar función de cálculo
            asistencia_id = asistencia_response.data[0]["id"]
            supabase_client.rpc("calcular_estado_asistencia", {"p_asistencia_id": asistencia_id}).execute()
        
        # Obtener nombre de quien registró
        registrado_por_nombre = None
        if marca.registrado_por:
            user_response = supabase_client.from_("usuario_sistema") \
                .select("nombre, apellido_paterno") \
                .eq("id", marca.registrado_por) \
                .execute()
            if user_response.data:
                user = user_response.data[0]
                registrado_por_nombre = f"{user['nombre']} {user.get('apellido_paterno', '')}".strip()
        
        return MarcaAsistenciaResponse(
            **marca_creada,
            registrado_por_nombre=registrado_por_nombre
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al registrar marca: {str(e)}")


# ============================================================================
# ENDPOINT 3: JUSTIFICAR AUSENCIA/ATRASO
# ============================================================================

@attendance_router.post("/asistencia/{asistencia_id}/justificar", response_model=EstadoAsistenciaResponse)
async def justificar_ausencia(
    asistencia_id: int,
    justificacion: JustificacionCreate
):
    """
    📝 Justifica una ausencia o atraso.
    
    **Permisos:** Solo admin/RRHH
    **Efecto:** Cambia estado a JUSTIFICADO, registra auditoría
    """
    try:
        # Verificar que existe el registro de asistencia
        asistencia_response = supabase_client.from_("asistencia") \
            .select("id") \
            .eq("id", asistencia_id) \
            .execute()
        
        if not asistencia_response.data:
            raise HTTPException(status_code=404, detail="Registro de asistencia no encontrado")
        
        # Actualizar o insertar estado justificado
        estado_data = {
            "asistencia_id": asistencia_id,
            "estado": "JUSTIFICADO",
            "tipo_justificacion": justificacion.tipo_justificacion.value,
            "justificacion": justificacion.justificacion,
            "justificado_por": justificacion.justificado_por,
            "fecha_justificacion": datetime.now().isoformat()
        }
        
        # Verificar si ya existe un estado
        existing_response = supabase_client.from_("asistencia_estados") \
            .select("id") \
            .eq("asistencia_id", asistencia_id) \
            .execute()
        
        if existing_response.data:
            # Actualizar
            response = supabase_client.from_("asistencia_estados") \
                .update(estado_data) \
                .eq("asistencia_id", asistencia_id) \
                .execute()
        else:
            # Insertar
            estado_data["minutos_atraso"] = 0
            response = supabase_client.from_("asistencia_estados").insert(estado_data).execute()
        
        if not response.data:
            raise HTTPException(status_code=500, detail="No se pudo registrar la justificación")
        
        return EstadoAsistenciaResponse(**response.data[0])
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al justificar ausencia: {str(e)}")


# ============================================================================
# ENDPOINT 4: HISTORIAL DE ASISTENCIA DE UN DOCTOR
# ============================================================================

@attendance_router.get("/doctor/{doctor_id}/historial")
async def obtener_historial_doctor(
    doctor_id: int,
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    limit: int = Query(30, le=100)
):
    """
    📜 Historial de asistencia de un doctor específico.
    
    **Usado en:** Panel lateral de detalle, perfil de doctor
    """
    try:
        # Fechas por defecto: últimos 30 días
        if not fecha_desde:
            fecha_desde = date.today() - timedelta(days=30)
        if not fecha_hasta:
            fecha_hasta = date.today()
        
        # Consultar historial
        response = supabase_client.from_("v_asistencia_consolidada") \
            .select("*") \
            .eq("usuario_sistema_id", doctor_id) \
            .gte("inicio_turno", f"{fecha_desde}T00:00:00") \
            .lte("inicio_turno", f"{fecha_hasta}T23:59:59") \
            .order("inicio_turno", desc=True) \
            .limit(limit) \
            .execute()
        
        return {
            "doctor_id": doctor_id,
            "fecha_desde": fecha_desde,
            "fecha_hasta": fecha_hasta,
            "total_registros": len(response.data) if response.data else 0,
            "historial": response.data or []
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener historial: {str(e)}")


# ============================================================================
# ENDPOINT 5: ESTADÍSTICAS DE ASISTENCIA
# ============================================================================

@attendance_router.get("/doctor/{doctor_id}/estadisticas", response_model=EstadisticasAsistenciaDoctor)
async def obtener_estadisticas_doctor(
    doctor_id: int,
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None)
):
    """
    📈 Estadísticas agregadas de asistencia de un doctor.
    
    **Métricas:**
    - Puntualidad, asistencia, atrasos, ausencias
    - Promedio de minutos trabajados
    - Pacientes atendidos en el periodo
    """
    try:
        # Fechas por defecto: mes actual
        if not fecha_desde:
            fecha_desde = date.today().replace(day=1)
        if not fecha_hasta:
            fecha_hasta = date.today()
        
        # Obtener datos del doctor
        doctor_response = supabase_client.from_("usuario_sistema") \
            .select("nombre, apellido_paterno") \
            .eq("id", doctor_id) \
            .execute()
        
        if not doctor_response.data:
            raise HTTPException(status_code=404, detail="Doctor no encontrado")
        
        doctor = doctor_response.data[0]
        doctor_nombre = f"{doctor['nombre']} {doctor.get('apellido_paterno', '')}".strip()
        
        # Consultar asistencias del periodo
        response = supabase_client.from_("v_asistencia_consolidada") \
            .select("*") \
            .eq("usuario_sistema_id", doctor_id) \
            .gte("inicio_turno", f"{fecha_desde}T00:00:00") \
            .lte("inicio_turno", f"{fecha_hasta}T23:59:59") \
            .execute()
        
        registros = response.data or []
        total_turnos = len(registros)
        
        if total_turnos == 0:
            return EstadisticasAsistenciaDoctor(
                doctor_id=doctor_id,
                doctor_nombre=doctor_nombre,
                periodo_inicio=fecha_desde,
                periodo_fin=fecha_hasta,
                total_turnos=0,
                dias_asistio=0,
                dias_atraso=0,
                dias_ausente=0,
                dias_justificado=0,
                promedio_minutos_atraso=0.0,
                porcentaje_puntualidad=0.0,
                porcentaje_asistencia=0.0,
                total_minutos_trabajados=0,
                total_pacientes_atendidos=0
            )
        
        # Calcular métricas
        dias_asistio = len([r for r in registros if r.get("estado") == "ASISTIO"])
        dias_atraso = len([r for r in registros if r.get("estado") == "ATRASO"])
        dias_ausente = len([r for r in registros if r.get("estado") == "AUSENTE"])
        dias_justificado = len([r for r in registros if r.get("estado") == "JUSTIFICADO"])
        
        minutos_atraso_total = sum([r.get("minutos_atraso", 0) for r in registros])
        promedio_minutos_atraso = minutos_atraso_total / total_turnos
        
        total_minutos_trabajados = sum([r.get("minutos_trabajados", 0) for r in registros if r.get("minutos_trabajados")])
        
        # Porcentajes
        porcentaje_puntualidad = ((dias_asistio + dias_justificado) / total_turnos) * 100
        porcentaje_asistencia = ((total_turnos - dias_ausente) / total_turnos) * 100
        
        # Pacientes atendidos
        pacientes_response = supabase_client.from_("cita_medica") \
            .select("id", count="exact") \
            .eq("usuario_sistema_id", doctor_id) \
            .eq("estado", "COMPLETADA") \
            .gte("fecha_cita", f"{fecha_desde}T00:00:00") \
            .lte("fecha_cita", f"{fecha_hasta}T23:59:59") \
            .execute()
        
        total_pacientes_atendidos = pacientes_response.count or 0
        
        return EstadisticasAsistenciaDoctor(
            doctor_id=doctor_id,
            doctor_nombre=doctor_nombre,
            periodo_inicio=fecha_desde,
            periodo_fin=fecha_hasta,
            total_turnos=total_turnos,
            dias_asistio=dias_asistio,
            dias_atraso=dias_atraso,
            dias_ausente=dias_ausente,
            dias_justificado=dias_justificado,
            promedio_minutos_atraso=round(promedio_minutos_atraso, 2),
            porcentaje_puntualidad=round(porcentaje_puntualidad, 2),
            porcentaje_asistencia=round(porcentaje_asistencia, 2),
            total_minutos_trabajados=total_minutos_trabajados,
            total_pacientes_atendidos=total_pacientes_atendidos
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al calcular estadísticas: {str(e)}")


# ============================================================================
# ENDPOINT 6: OBTENER PARÁMETROS DE CONFIGURACIÓN
# ============================================================================

@attendance_router.get("/parametros", response_model=List[ParametroAsistencia])
async def obtener_parametros():
    """
    ⚙️ Obtiene los parámetros configurables del módulo de asistencia.
    
    **Usado en:** Página de ajustes del sistema
    """
    try:
        response = supabase_client.from_("parametros_asistencia") \
            .select("*") \
            .eq("activo", True) \
            .order("parametro") \
            .execute()
        
        return [ParametroAsistencia(**param) for param in response.data] if response.data else []
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener parámetros: {str(e)}")


# ============================================================================
# ENDPOINT 7: ACTUALIZAR PARÁMETRO
# ============================================================================

@attendance_router.patch("/parametros/{parametro_id}", response_model=ParametroAsistencia)
async def actualizar_parametro(
    parametro_id: int,
    actualizacion: ParametroAsistenciaUpdate
):
    """
    🔧 Actualiza un parámetro de configuración.
    
    **Permisos:** Solo admin
    """
    try:
        update_data = actualizacion.model_dump(exclude_unset=True)
        
        if not update_data:
            raise HTTPException(status_code=400, detail="No hay datos para actualizar")
        
        response = supabase_client.from_("parametros_asistencia") \
            .update(update_data) \
            .eq("id", parametro_id) \
            .execute()
        
        if not response.data:
            raise HTTPException(status_code=404, detail="Parámetro no encontrado")
        
        return ParametroAsistencia(**response.data[0])
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al actualizar parámetro: {str(e)}")


# ============================================================================
# ENDPOINT 8: OBTENER MARCAS DE UN TURNO ESPECÍFICO
# ============================================================================

@attendance_router.get("/asistencia/{asistencia_id}/marcas", response_model=List[MarcaAsistenciaResponse])
async def obtener_marcas_turno(asistencia_id: int):
    """
    🔍 Obtiene todas las marcas (entrada/salida) de un turno específico.
    
    **Usado en:** Panel lateral de detalle, auditoría
    """
    try:
        # Obtener fecha del turno
        asistencia_response = supabase_client.from_("asistencia") \
            .select("inicio_turno, usuario_sistema_id") \
            .eq("id", asistencia_id) \
            .execute()
        
        if not asistencia_response.data:
            raise HTTPException(status_code=404, detail="Registro de asistencia no encontrado")
        
        turno = asistencia_response.data[0]
        fecha_turno = turno["inicio_turno"][:10]
        
        # Obtener marcas del día para ese doctor
        marcas_response = supabase_client.from_("marcas_asistencia") \
            .select("*") \
            .eq("usuario_sistema_id", turno["usuario_sistema_id"]) \
            .gte("fecha_hora_marca", f"{fecha_turno}T00:00:00") \
            .lte("fecha_hora_marca", f"{fecha_turno}T23:59:59") \
            .order("fecha_hora_marca", desc=False) \
            .execute()
        
        marcas = []
        for marca in marcas_response.data or []:
            # Obtener nombre de quien registró
            registrado_por_nombre = None
            if marca.get("registrado_por"):
                user_response = supabase_client.from_("usuario_sistema") \
                    .select("nombre, apellido_paterno") \
                    .eq("id", marca["registrado_por"]) \
                    .execute()
                if user_response.data:
                    user = user_response.data[0]
                    registrado_por_nombre = f"{user['nombre']} {user.get('apellido_paterno', '')}".strip()
            
            marcas.append(MarcaAsistenciaResponse(
                **marca,
                registrado_por_nombre=registrado_por_nombre
            ))
        
        return marcas

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener marcas: {str(e)}")


# ============================================================================
# ENDPOINT 9: GENERAR REPORTE EXCEL
# ============================================================================

@attendance_router.get("/reporte/excel")
async def generar_reporte_excel(
    fecha: Optional[date] = Query(None, description="Fecha del reporte (default: hoy)")
):
    """
    📊 Genera un reporte Excel profesional de asistencia del día.

    **Incluye:**
    - Resumen ejecutivo con KPIs
    - Listado detallado de turnos
    - Gráficos de cumplimiento
    - Formato corporativo moderno
    """
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from fastapi.responses import StreamingResponse

        fecha_reporte = fecha or date.today()

        # Obtener datos del día
        response = supabase_client.from_("v_asistencia_consolidada") \
            .select("*") \
            .gte("inicio_turno", f"{fecha_reporte}T00:00:00") \
            .lte("inicio_turno", f"{fecha_reporte}T23:59:59") \
            .order("inicio_turno", desc=False) \
            .execute()

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencia"

        # Estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=14, color="1F4E78")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = f"REPORTE DE ASISTENCIA - {fecha_reporte.strftime('%d/%m/%Y')}"
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # KPIs
        ws.merge_cells('A3:B3')
        ws['A3'] = "RESUMEN DEL DÍA"
        ws['A3'].font = Font(bold=True, size=11)

        turnos = response.data or []
        stats = {
            "total": len(turnos),
            "en_turno": len([t for t in turnos if t.get("estado") == "EN_TURNO"]),
            "asistieron": len([t for t in turnos if t.get("estado") == "ASISTIO"]),
            "atrasos": len([t for t in turnos if t.get("estado") == "ATRASO"]),
            "ausentes": len([t for t in turnos if t.get("estado") == "AUSENTE"]),
            "justificados": len([t for t in turnos if t.get("estado") == "JUSTIFICADO"])
        }

        ws['A4'] = "Total Turnos:"
        ws['B4'] = stats["total"]
        ws['A5'] = "En Turno:"
        ws['B5'] = stats["en_turno"]
        ws['C4'] = "Asistieron:"
        ws['D4'] = stats["asistieron"]
        ws['C5'] = "Con Atraso:"
        ws['D5'] = stats["atrasos"]
        ws['E4'] = "Ausentes:"
        ws['F4'] = stats["ausentes"]
        ws['E5'] = "Justificados:"
        ws['F5'] = stats["justificados"]

        # Headers de la tabla
        row = 7
        headers = ["Doctor", "RUT", "Entrada Programada", "Entrada Real", "Salida Real",
                   "Tiempo Trabajado (min)", "Atraso (min)", "Estado"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Datos
        row = 8
        for turno in turnos:
            # Obtener datos del doctor
            doctor_response = supabase_client.from_("usuario_sistema") \
                .select("nombre, apellido_paterno, apellido_materno, rut") \
                .eq("id", turno["doctor_id"]) \
                .single() \
                .execute()

            doctor = doctor_response.data if doctor_response.data else {}
            doctor_nombre = f"{doctor.get('nombre', '')} {doctor.get('apellido_paterno', '')} {doctor.get('apellido_materno', '')}".strip()

            ws.cell(row=row, column=1, value=doctor_nombre).border = border
            ws.cell(row=row, column=2, value=doctor.get('rut', '')).border = border
            ws.cell(row=row, column=3, value=turno["inicio_turno"]).border = border
            ws.cell(row=row, column=4, value=turno.get("hora_entrada_real", "")).border = border
            ws.cell(row=row, column=5, value=turno.get("hora_salida_real", "")).border = border
            ws.cell(row=row, column=6, value=turno.get("minutos_trabajados", 0)).border = border
            ws.cell(row=row, column=7, value=turno.get("minutos_atraso", 0)).border = border

            estado_cell = ws.cell(row=row, column=8, value=turno.get("estado", "EN_TURNO"))
            estado_cell.border = border

            # Color según estado
            estado = turno.get("estado", "EN_TURNO")
            if estado == "ASISTIO":
                estado_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif estado == "ATRASO":
                estado_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif estado == "AUSENTE":
                estado_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif estado == "JUSTIFICADO":
                estado_cell.fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")

            row += 1

        # Ajustar anchos de columna
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=asistencia_{fecha_reporte}.xlsx"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar reporte Excel: {str(e)}")


# ============================================================================
# ENDPOINT 10: GENERAR REPORTE PDF
# ============================================================================

@attendance_router.get("/reporte/pdf")
async def generar_reporte_pdf(
    fecha: Optional[date] = Query(None, description="Fecha del reporte (default: hoy)")
):
    """
    📄 Genera un reporte PDF profesional de asistencia del día.

    **Incluye:**
    - Encabezado corporativo
    - KPIs visuales
    - Tabla detallada con colores
    - Pie de página con fecha de generación
    """
    try:
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from fastapi.responses import StreamingResponse

        fecha_reporte = fecha or date.today()

        # Obtener datos del día
        response = supabase_client.from_("v_asistencia_consolidada") \
            .select("*") \
            .gte("inicio_turno", f"{fecha_reporte}T00:00:00") \
            .lte("inicio_turno", f"{fecha_reporte}T23:59:59") \
            .order("inicio_turno", desc=False) \
            .execute()

        turnos = response.data or []

        # Calcular estadísticas
        stats = {
            "total": len(turnos),
            "en_turno": len([t for t in turnos if t.get("estado") == "EN_TURNO"]),
            "asistieron": len([t for t in turnos if t.get("estado") == "ASISTIO"]),
            "atrasos": len([t for t in turnos if t.get("estado") == "ATRASO"]),
            "ausentes": len([t for t in turnos if t.get("estado") == "AUSENTE"]),
            "justificados": len([t for t in turnos if t.get("estado") == "JUSTIFICADO"])
        }

        # Crear PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1F4E78'),
            alignment=TA_CENTER,
            spaceAfter=20
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=10
        )

        # Título
        elements.append(Paragraph(f"REPORTE DE ASISTENCIA", title_style))
        elements.append(Paragraph(f"Fecha: {fecha_reporte.strftime('%d de %B de %Y')}", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))

        # KPIs
        elements.append(Paragraph("RESUMEN DEL DÍA", subtitle_style))

        kpi_data = [
            ['Total Turnos', 'En Turno', 'Asistieron', 'Con Atraso', 'Ausentes', 'Justificados'],
            [str(stats['total']), str(stats['en_turno']), str(stats['asistieron']),
             str(stats['atrasos']), str(stats['ausentes']), str(stats['justificados'])]
        ]

        kpi_table = Table(kpi_data, colWidths=[1.2*inch]*6)
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
        ]))

        elements.append(kpi_table)
        elements.append(Spacer(1, 0.3*inch))

        # Tabla de detalles
        elements.append(Paragraph("DETALLE DE ASISTENCIA", subtitle_style))

        table_data = [['Doctor', 'RUT', 'Entrada\nProgramada', 'Entrada\nReal', 'Salida\nReal',
                       'Tiempo\nTrabajado', 'Atraso\n(min)', 'Estado']]

        for turno in turnos:
            # Obtener datos del doctor
            doctor_response = supabase_client.from_("usuario_sistema") \
                .select("nombre, apellido_paterno, apellido_materno, rut") \
                .eq("id", turno["doctor_id"]) \
                .single() \
                .execute()

            doctor = doctor_response.data if doctor_response.data else {}
            doctor_nombre = f"{doctor.get('nombre', '')} {doctor.get('apellido_paterno', '')}".strip()

            entrada_prog = datetime.fromisoformat(turno["inicio_turno"].replace("Z", "+00:00")).strftime("%H:%M")
            entrada_real = datetime.fromisoformat(turno["hora_entrada_real"].replace("Z", "+00:00")).strftime("%H:%M") if turno.get("hora_entrada_real") else "--:--"
            salida_real = datetime.fromisoformat(turno["hora_salida_real"].replace("Z", "+00:00")).strftime("%H:%M") if turno.get("hora_salida_real") else "--:--"

            minutos = turno.get("minutos_trabajados", 0)
            tiempo_trabajado = f"{minutos // 60}h {minutos % 60}m" if minutos else "--"

            table_data.append([
                doctor_nombre,
                doctor.get('rut', '')[:12],
                entrada_prog,
                entrada_real,
                salida_real,
                tiempo_trabajado,
                str(turno.get("minutos_atraso", 0)),
                turno.get("estado", "EN_TURNO")
            ])

        detail_table = Table(table_data, colWidths=[1.3*inch, 0.9*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.9*inch, 0.6*inch, 0.8*inch])

        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]

        # Aplicar colores por estado
        for i, turno in enumerate(turnos, start=1):
            estado = turno.get("estado", "EN_TURNO")
            if estado == "ASISTIO":
                table_style.append(('BACKGROUND', (7, i), (7, i), colors.HexColor('#C6EFCE')))
            elif estado == "ATRASO":
                table_style.append(('BACKGROUND', (7, i), (7, i), colors.HexColor('#FFEB9C')))
            elif estado == "AUSENTE":
                table_style.append(('BACKGROUND', (7, i), (7, i), colors.HexColor('#FFC7CE')))
            elif estado == "JUSTIFICADO":
                table_style.append(('BACKGROUND', (7, i), (7, i), colors.HexColor('#E4DFEC')))

        detail_table.setStyle(TableStyle(table_style))
        elements.append(detail_table)

        # Pie de página
        elements.append(Spacer(1, 0.5*inch))
        footer_text = f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')} | Sistema de Gestión Clínica"
        elements.append(Paragraph(footer_text, ParagraphStyle('Footer', parent=styles['Normal'],
                                                               fontSize=8, textColor=colors.grey,
                                                               alignment=TA_CENTER)))

        # Construir PDF
        doc.build(elements)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=asistencia_{fecha_reporte}.pdf"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar reporte PDF: {str(e)}")

